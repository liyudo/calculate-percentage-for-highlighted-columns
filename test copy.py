import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Open workbook and select sheet
wb = openpyxl.load_workbook('building.xlsx')
ws = wb.active

# Create new sheet and set header row
ws_percent = wb.create_sheet('percentage')
header_row = []
highlighted_cols = []
for cell in ws[1]:
    if cell.fill.fgColor.rgb == 'FFFFFF00':
        highlighted_cols.append(cell.column)
if not highlighted_cols:
    print("Error: No highlighted columns found.")
    exit()

# Iterate over columns and calculate percentages
for col_idx, col in enumerate(highlighted_cols):
    column_data = [cell.value for cell in ws[get_column_letter(col)][1:] if cell.value is not None]
    if not column_data:
        print(f"Warning: No data found in column {get_column_letter(col)}")
        continue
    unique_values = set(column_data)
    count_dict = {value: column_data.count(value) for value in unique_values}
    total_count = len(column_data)
    header = ws[get_column_letter(col) + "1"].value
    ws_percent.cell(row=1, column=col_idx*2+1, value=header)
    ws_percent.cell(row=1, column=col_idx*2+2, value=f"{header}_percentage")
    for i, value in enumerate(unique_values):
        ws_percent.cell(row=i+2, column=col_idx*2+1, value=value)
        count = count_dict[value]
        percentage = count / total_count
        ws_percent.cell(row=i+2, column=col_idx*2+2, value=percentage)
        ws_percent.cell(row=i+2, column=col_idx*2+2).number_format = '0.00%'
    
# Save workbook
wb.save('building.xlsx')
